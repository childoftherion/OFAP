#!/usr/bin/env python3
"""
Enhanced analysis of Oregon Marijuana Tax Receipts and Revenue Distributions
Properly extracts all columns and revenue figures for year-over-year and monthly analysis
"""

import json
import sys
from pathlib import Path
from datetime import datetime
from decimal import Decimal

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Install with: pip3 install --user --break-system-packages openpyxl")
    sys.exit(1)

def parse_value(value):
    """Parse cell value, handling dates, numbers, and strings"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, str):
        # Try to parse as number
        value = value.strip().replace('$', '').replace(',', '')
        if value == '' or value == '-':
            return None
        try:
            return float(value)
        except:
            return value
    return str(value)

def analyze_tax_receipts(file_path):
    """Analyze monthly marijuana tax receipts with proper column extraction"""
    print(f"\n=== Analyzing Tax Receipts: {file_path.name} ===\n")
    
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    print(f"Sheet: {ws.title}")
    print(f"Rows: {ws.max_row}, Columns: {ws.max_column}\n")
    
    # Find actual header row (look for "Month" or date values)
    header_row = None
    for row in range(1, min(20, ws.max_row + 1)):
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value and ('Month' in str(cell_value) or 'Date' in str(cell_value)):
                header_row = row
                break
        if header_row:
            break
    
    if not header_row:
        header_row = 1
    
    # Read headers
    headers = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        headers.append(parse_value(cell.value) if cell.value else f"Column{col}")
    
    print(f"Headers: {headers}\n")
    
    # Extract data - start after header row
    data = []
    for row in range(header_row + 1, ws.max_row + 1):
        row_data = {}
        has_value = False
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            value = parse_value(cell.value)
            if value is not None and value != '':
                has_value = True
            row_data[header] = value
        if has_value:
            data.append(row_data)
    
    print(f"Extracted {len(data)} data rows\n")
    
    # Calculate statistics
    monthly_totals = {}
    yearly_totals = {}
    
    # Find date and amount columns
    date_col = None
    amount_cols = []
    for i, header in enumerate(headers):
        header_str = str(header).lower()
        if 'month' in header_str or 'date' in header_str:
            date_col = i
        elif any(keyword in header_str for keyword in ['state', 'local', 'total', 'amount', 'receipt']):
            if 'note' not in header_str:
                amount_cols.append(i)
    
    # Calculate monthly and yearly totals
    for row_data in data:
        if date_col is not None and headers[date_col] in row_data:
            date_val = row_data[headers[date_col]]
            if date_val:
                # Extract year-month
                if isinstance(date_val, str):
                    try:
                        # Try parsing date string
                        from dateutil import parser
                        dt = parser.parse(date_val)
                    except:
                        dt = None
                elif isinstance(date_val, datetime):
                    dt = date_val
                else:
                    dt = None
                
                if dt:
                    year_month = f"{dt.year}-{dt.month:02d}"
                    year = dt.year
                    
                    # Sum amounts
                    for col_idx in amount_cols:
                        if headers[col_idx] in row_data:
                            amount = row_data[headers[col_idx]]
                            if isinstance(amount, (int, float)) and amount > 0:
                                monthly_totals[year_month] = monthly_totals.get(year_month, 0) + amount
                                yearly_totals[year] = yearly_totals.get(year, 0) + amount
    
    return {
        'headers': [str(h) for h in headers],
        'data': data[:100],  # First 100 rows
        'total_rows': len(data),
        'monthly_totals': {k: float(v) for k, v in monthly_totals.items()},
        'yearly_totals': {k: float(v) for k, v in yearly_totals.items()},
        'date_column_index': date_col,
        'amount_column_indices': amount_cols
    }

def analyze_revenue_distributions(file_path):
    """Analyze quarterly revenue distributions with proper column extraction"""
    print(f"\n=== Analyzing Revenue Distributions: {file_path.name} ===\n")
    
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    print(f"Sheet: {ws.title}")
    print(f"Rows: {ws.max_row}, Columns: {ws.max_column}\n")
    
    # Find header row
    header_row = None
    for row in range(1, min(20, ws.max_row + 1)):
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value and ('Quarter' in str(cell_value) or 'Q' in str(cell_value)):
                header_row = row
                break
        if header_row:
            break
    
    if not header_row:
        header_row = 1
    
    # Read headers
    headers = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        headers.append(parse_value(cell.value) if cell.value else f"Column{col}")
    
    print(f"Headers: {headers}\n")
    
    # Extract data
    data = []
    quarterly_totals = {}
    for row in range(header_row + 1, ws.max_row + 1):
        row_data = {}
        has_value = False
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            value = parse_value(cell.value)
            if value is not None and value != '':
                has_value = True
            row_data[header] = value
        if has_value:
            data.append(row_data)
            # Try to find quarter identifier and total
            for header, val in row_data.items():
                if val and isinstance(val, str) and ('Q' in str(val) or 'quarter' in str(val).lower()):
                    # Look for amount columns
                    for amt_header, amt_val in row_data.items():
                        if isinstance(amt_val, (int, float)) and amt_val > 0:
                            quarterly_totals[str(val)] = quarterly_totals.get(str(val), 0) + amt_val
    
    print(f"Extracted {len(data)} data rows\n")
    print(f"Quarterly totals found: {len(quarterly_totals)} quarters\n")
    
    return {
        'headers': [str(h) for h in headers],
        'data': data[:100],
        'total_rows': len(data),
        'quarterly_totals': {k: float(v) for k, v in quarterly_totals.items()}
    }

def main():
    base_path = Path(__file__).parent
    
    # Analyze tax receipts
    receipts_xlsx = base_path / "marijuana-tax-receipts" / "marijuana-tax-receipts-most-recent.xlsx"
    receipts_data = analyze_tax_receipts(receipts_xlsx)
    
    # Analyze distributions
    dist_xlsx = base_path / "revenue-distributions" / "revenue-distributions-most-recent.xlsx"
    dist_data = analyze_revenue_distributions(dist_xlsx)
    
    # Save detailed analysis
    analysis = {
        'timestamp': datetime.now().isoformat(),
        'source_files': {
            'receipts': str(receipts_xlsx),
            'distributions': str(dist_xlsx)
        },
        'tax_receipts': receipts_data,
        'revenue_distributions': dist_data
    }
    
    output_file = base_path / "analysis_results_detailed.json"
    with open(output_file, 'w') as f:
        json.dump(analysis, f, indent=2, default=str)
    
    print(f"\n=== Analysis complete ===")
    print(f"Results saved to: {output_file}")
    
    # Print summary
    print("\n=== SUMMARY ===")
    if receipts_data.get('yearly_totals'):
        print("\nYearly Tax Receipts:")
        for year in sorted(receipts_data['yearly_totals'].keys()):
            print(f"  {year}: ${receipts_data['yearly_totals'][year]:,.2f}")
    
    if receipts_data.get('monthly_totals'):
        print(f"\nMonthly totals available for {len(receipts_data['monthly_totals'])} months")
    
    return analysis

if __name__ == '__main__':
    try:
        from dateutil import parser
    except ImportError:
        print("Installing python-dateutil...")
        import subprocess
        subprocess.run([sys.executable, '-m', 'pip', 'install', '--user', '--break-system-packages', 'python-dateutil'], check=True)
        from dateutil import parser
    
    main()

