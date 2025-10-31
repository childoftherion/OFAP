#!/usr/bin/env python3
"""
Analyze Oregon Marijuana Tax Receipts and Revenue Distributions
Extracts data from DOR XLSX files and performs year-over-year and monthly analysis
"""

import json
import sys
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Install with: pip3 install --user openpyxl")
    sys.exit(1)

def analyze_tax_receipts(file_path):
    """Analyze monthly marijuana tax receipts"""
    print(f"\n=== Analyzing Tax Receipts: {file_path.name} ===\n")
    
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # Print sheet info
        print(f"Sheet: {ws.title}")
        print(f"Rows: {ws.max_row}, Columns: {ws.max_column}\n")
        
        # Read header row
        headers = []
        header_row = 1
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col)
            if cell.value:
                headers.append(str(cell.value).strip())
        
        print(f"Headers: {headers}\n")
        
        # Extract data rows
        data = []
        for row in range(2, min(ws.max_row + 1, 100)):  # Limit to first 100 rows
            row_data = {}
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value is not None:
                    row_data[header] = cell.value
            if row_data:
                data.append(row_data)
        
        return {
            'headers': headers,
            'data': data[:50],  # First 50 rows
            'total_rows': len(data)
        }
        
    except Exception as e:
        return {'error': str(e)}

def analyze_revenue_distributions(file_path):
    """Analyze quarterly revenue distributions"""
    print(f"\n=== Analyzing Revenue Distributions: {file_path.name} ===\n")
    
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        print(f"Sheet: {ws.title}")
        print(f"Rows: {ws.max_row}, Columns: {ws.max_column}\n")
        
        # Read header row
        headers = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            if cell.value:
                headers.append(str(cell.value).strip())
        
        print(f"Headers: {headers}\n")
        
        # Extract data rows
        data = []
        for row in range(2, min(ws.max_row + 1, 100)):
            row_data = {}
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value is not None:
                    row_data[header] = cell.value
            if row_data:
                data.append(row_data)
        
        return {
            'headers': headers,
            'data': data[:50],
            'total_rows': len(data)
        }
        
    except Exception as e:
        return {'error': str(e)}

def main():
    base_path = Path(__file__).parent
    
    # Analyze tax receipts
    receipts_xlsx = base_path / "marijuana-tax-receipts" / "marijuana-tax-receipts-most-recent.xlsx"
    receipts_data = analyze_tax_receipts(receipts_xlsx)
    
    # Analyze distributions
    dist_xlsx = base_path / "revenue-distributions" / "revenue-distributions-most-recent.xlsx"
    dist_data = analyze_revenue_distributions(dist_xlsx)
    
    # Save analysis results
    analysis = {
        'timestamp': datetime.now().isoformat(),
        'tax_receipts': receipts_data,
        'revenue_distributions': dist_data
    }
    
    output_file = base_path / "analysis_results.json"
    with open(output_file, 'w') as f:
        json.dump(analysis, f, indent=2, default=str)
    
    print(f"\n=== Analysis complete ===")
    print(f"Results saved to: {output_file}")
    
    return analysis

if __name__ == '__main__':
    main()

